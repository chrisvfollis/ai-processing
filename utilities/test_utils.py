# standard dependencies
import os
import re
from typing import Union
import pickle

# 3rd-party dependencies
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import boto3


# internal dependencies
from utilities import io_utils


def add_imgs_to_spreadsheet(df):
    csv_path = 'output/face_data.csv'
    df = pd.read_csv(csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Face Data"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    image_column = len(df.columns) + 1
    ws.cell(row=1, column=image_column, value="image")

    for idx, row in df.iterrows():
        img_path = f'output/faces/{idx}.jpg' 
        if os.path.exists(img_path):
            img = XLImage(img_path)
            img.height = 80
            img.width = 80
            ws.add_image(img, f"{chr(65 + image_column - 1)}{idx + 2}")

    output_path = 'output/face_data.xlsx'
    wb.save(output_path)

    output_path


def download_tracking_pkls(
        shop_id: str = None,
        bucket_name='visionservice-data',
        local_dir='../files/output/'
    ):
    credentials = io_utils.get_aws_creds()

    s3 = boto3.client(
        's3',
        aws_access_key_id=credentials[0],
        aws_secret_access_key=credentials[1],
        region_name='us-west-1'
    )

    os.makedirs(local_dir, exist_ok=True)

    pattern = re.compile(r'\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}_\d+\.pkl')
    # pattern match example: 2025-04-18_09-30-00_3.pkl

    prefix = f'{shop_id}/' if shop_id else ''

    paginator = s3.get_paginator('list_objects_v2')
    page_iterator = paginator.paginate(Bucket=bucket_name, Prefix=prefix)

    for page in page_iterator:
        if 'Contents' not in page:
            continue

        for obj in page['Contents']:
            key = obj['Key']
            filename = os.path.basename(key)

            if pattern.fullmatch(filename):
                local_path = os.path.join(local_dir, filename)
                if os.path.exists(local_path):
                    continue
                print(f'Downloading {key} to {local_path}')
                s3.download_file(bucket_name, key, local_path)


def prepare_tracking_data(
        pkl_dir='../files/output/',
        min_duration_sec: Union[float, int] = 1.0,
        var_percentile: int = 5
    ):
    '''
    Loads and cleans pickled tracking data.

    Args:
        min_duration_sec: Filters short-lived tracks likely caused by transient
                          object detector noise. Tracks with durations below this
                          threshold are discarded.

        var_percentile: Filters stationary objects erroneously tracked as people
                        for extended periods. A track's average [x, y, w, h] variance
                        must fall above this percentile within the distribution of
                        all track-wise variances.

    Returns:
        List of (fps, cleaned_tracks) tuples.
    '''

    all_variances = []
    loaded_data = []

    len_filtered = 0
    var_filtered = 0

    # load and compute variances
    for fname in os.listdir(pkl_dir):
        if fname.endswith('.pkl') and not fname.endswith('inference_data.pkl'):
            path = os.path.join(pkl_dir, fname)
            try:
                with open(path, 'rb') as f:
                    pipeline = pickle.load(f)

                fps = pipeline.fps
                track_variances = []

                for trk in pipeline.all_trks.values():
                    if not hasattr(trk, 'span') or not isinstance(trk.span, list):
                        continue
                    start, end = trk.span
                    duration = max(0, end - start) / fps
                    if duration <= min_duration_sec:
                        len_filtered += 1
                        continue  # filter short, irrelevant tracks

                    boxes = list(trk.object_detections.values())
                    if len(boxes) < 2:
                        continue

                    arr = np.array([[box[0], box[1], box[2], box[3]]
                                    for box in boxes if len(box) >= 4])
                    if arr.shape[0] < 2:
                        continue

                    var = np.var(arr, axis=0).mean()
                    track_variances.append((trk, var))

                if track_variances:
                    loaded_data.append((fps, track_variances))
                    all_variances.extend([v for _, v in track_variances])

            except Exception as e:
                print(f'Failed to process {fname}: {e}')

    if not all_variances:
        return []
    
    # filter using variance threshold
    var_cutoff = np.percentile(all_variances, var_percentile)
    cleaned_data = []
    for fps, trk_var_pairs in loaded_data:
        cleaned_tracks = {}
        for trk, var in trk_var_pairs:
            if var > var_cutoff:
                cleaned_tracks[trk.track_id] = trk
            else:
                var_filtered += 1
        if cleaned_tracks:
            cleaned_data.append((fps, cleaned_tracks))
    
    print(f'{len_filtered} short-lived tracks filtered')
    print(f'{var_filtered} tracked stationary objects filtered')

    return cleaned_data


def export_tracking_analysis(
        all_trackwise_stats: list[pd.DataFrame] = None,
        all_overall_stats: list[pd.DataFrame] = None,
        all_ols_output: list[pd.DataFrame] = None
    ):

    merged_trackwise = (
        pd.concat(all_trackwise_stats, ignore_index=True)
        if all_trackwise_stats else pd.DataFrame()
    )
    merged_overall = (
        pd.concat(all_overall_stats, ignore_index=True)
        if all_overall_stats else pd.DataFrame()
    )
    merged_ols = (
        pd.concat(all_ols_output, ignore_index=True)
        if all_ols_output else pd.DataFrame()
    )

    filename = io_utils.get_unique_filename('../files/output/', 'tracking_analysis_results.xlsx')

    output_path = os.path.join('../files/output/', filename)
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        if not merged_trackwise.empty:
            merged_trackwise.to_excel(writer, sheet_name='Trackwise Stats', index=False)
        if not merged_overall.empty:
            merged_overall.to_excel(writer, sheet_name='Overall Stats', index=False)
        if not merged_ols.empty:
            merged_ols.to_excel(writer, sheet_name='OLS Models', index=False)

    print(f'Exported results to: {output_path}')
