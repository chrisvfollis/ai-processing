# standard dependencies
import os
import re
from typing import Union
import pickle
from datetime import datetime

# 3rd-party dependencies
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
import boto3

# internal dependencies
from utilities import io_utils


def add_imgs_to_spreadsheet(df):
    csv_path = 'output/face_data.csv'
    df = pd.read_csv(csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Face Data"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    image_column = len(df.columns) + 1
    ws.cell(row=1, column=image_column, value="image")

    for idx, row in df.iterrows():
        img_path = f'output/faces/{idx}.jpg' 
        if os.path.exists(img_path):
            img = XLImage(img_path)
            img.height = 160
            img.width = 160
            ws.add_image(img, f"{chr(65 + image_column - 1)}{idx + 2}")

    output_path = 'output/face_data.xlsx'
    wb.save(output_path)

    output_path


def download_tracking_pkls(
        shop_id: str = None,
        bucket_name='visionservice-data',
        local_dir='../files/output/'
    ):
    credentials = io_utils.get_aws_credentials()

    s3 = boto3.client(
        's3',
        aws_access_key_id=credentials[0],
        aws_secret_access_key=credentials[1],
        region_name='us-west-1'
    )

    os.makedirs(local_dir, exist_ok=True)

    pattern = re.compile(r'\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}_\d+\.pkl')
    # pattern match example: 2025-04-18_09-30-00_3.pkl

    prefix = f'{shop_id}/' if shop_id else ''

    paginator = s3.get_paginator('list_objects_v2')
    page_iterator = paginator.paginate(Bucket=bucket_name, Prefix=prefix)

    for page in page_iterator:
        if 'Contents' not in page:
            continue

        for obj in page['Contents']:
            key = obj['Key']
            filename = os.path.basename(key)

            if pattern.fullmatch(filename):
                local_path = os.path.join(local_dir, filename)
                if os.path.exists(local_path):
                    continue
                print(f'Downloading {key} to {local_path}')
                s3.download_file(bucket_name, key, local_path)


def download_event_imgs(
        bucket_name='timemanager-event-imgs',
        local_dir='../files/output/event_imgs',
        max_imgs=1000,
        start_from: Union[datetime, list] = None,
        min_bytes: int = 0
    ):

    print('Downloading event images...')
    credentials = io_utils.get_aws_credentials()
    s3 = boto3.client(
        's3',
        aws_access_key_id=credentials[0],
        aws_secret_access_key=credentials[1],
        region_name='us-west-1'
    )

    os.makedirs(local_dir, exist_ok=True)

    if isinstance(start_from, list):
        start_from = datetime(*start_from).replace(tzinfo=None)

    paginator = s3.get_paginator('list_objects_v2')
    page_iterator = paginator.paginate(Bucket=bucket_name)

    for page in page_iterator:
        if 'Contents' not in page:
            continue

        for obj in page['Contents']:
            num_local_imgs = len(os.listdir(local_dir))
            if num_local_imgs >= max_imgs:
                return
            
            last_modified = obj['LastModified'].replace(tzinfo=None)
            if start_from and (last_modified < start_from):
                continue

            if obj['Size'] < min_bytes:
                continue
    
            key = obj['Key']
            filename = os.path.basename(key)
            if not filename:  # skip keys that end in '/' (folders)
                continue

            local_path = os.path.join(local_dir, filename)
            if os.path.exists(local_path):
                continue

            s3.download_file(bucket_name, key, local_path)


def prepare_tracking_data(
        pkl_dir='../files/output/',
        min_duration_sec: Union[float, int] = 1.0,
        var_percentile: int = 5
    ):
    '''
    Loads and cleans pickled tracking data.

    Args:
        min_duration_sec: Filters short-lived tracks likely caused by transient
                          object detector noise. Tracks with durations below this
                          threshold are discarded.

        var_percentile: Filters stationary objects erroneously tracked as people
                        for extended periods. A track's average [x, y, w, h] variance
                        must fall above this percentile within the distribution of
                        all track-wise variances.

    Returns:
        List of (fps, cleaned_tracks) tuples.
    '''

    all_variances = []
    loaded_data = []

    len_filtered = 0
    var_filtered = 0

    # load and compute variances
    for fname in os.listdir(pkl_dir):
        if fname.endswith('.pkl') and not fname.endswith('inference_data.pkl'):
            path = os.path.join(pkl_dir, fname)
            try:
                with open(path, 'rb') as f:
                    pipeline = pickle.load(f)

                fps = pipeline.fps
                track_variances = []

                for trk in pipeline.all_trks.values():
                    if not hasattr(trk, 'span') or not isinstance(trk.span, list):
                        continue
                    start, end = trk.span
                    duration = max(0, end - start) / fps
                    if duration <= min_duration_sec:
                        len_filtered += 1
                        continue  # filter short, irrelevant tracks

                    boxes = list(trk.object_detections.values())
                    if len(boxes) < 2:
                        continue

                    arr = np.array([[box[0], box[1], box[2], box[3]]
                                    for box in boxes if len(box) >= 4])
                    if arr.shape[0] < 2:
                        continue

                    var = np.var(arr, axis=0).mean()
                    track_variances.append((trk, var))

                if track_variances:
                    loaded_data.append((fps, track_variances))
                    all_variances.extend([v for _, v in track_variances])

            except Exception as e:
                print(f'Failed to process {fname}: {e}')

    if not all_variances:
        return []
    
    # filter using variance threshold
    var_cutoff = np.percentile(all_variances, var_percentile)
    cleaned_data = []
    for fps, trk_var_pairs in loaded_data:
        cleaned_tracks = {}
        for trk, var in trk_var_pairs:
            if var > var_cutoff:
                cleaned_tracks[trk.track_id] = trk
            else:
                var_filtered += 1
        if cleaned_tracks:
            cleaned_data.append((fps, cleaned_tracks))
    
    print(f'{len_filtered} short-lived tracks filtered')
    print(f'{var_filtered} tracked stationary objects filtered')

    return cleaned_data


def export_tracking_analysis(
        trackwise_stats: pd.DataFrame = None,
        overall_stats: pd.DataFrame = None,
        ols_output: pd.DataFrame = None
    ):


    filename = io_utils.get_unique_filename('../files/output/', 'tracking_analysis_results.xlsx')

    output_path = os.path.join('../files/output/', filename)
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        if not trackwise_stats.empty:
            trackwise_stats.to_excel(writer, sheet_name='Trackwise Stats', index=False)
        if not overall_stats.empty:
            overall_stats.to_excel(writer, sheet_name='Overall Stats', index=False)
        if not ols_output.empty:
            ols_output.to_excel(writer, sheet_name='OLS Models', index=False)

    print(f'Exported results to: {output_path}')


def export_face_event_spreadsheet(
        full_face_df, excel_path='../files/output/event_img_face_data.xlsx'
    ):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Event Image Face Data'

    image_paths = pd.DataFrame(full_face_df['img_path'])

    full_face_df = full_face_df[['img_area', 'face_area', 'distance', 'name']]
    columns = [col for col in full_face_df.columns]

    header = columns + ['correct_id', 'image']
    ws.append(header)

    # format header:
    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    excel_row_idx = 2  # start after header
    image_paths_to_remove = []

    ws.sheet_format.defaultRowHeight = 160

    for i, (_, row) in enumerate(full_face_df.iterrows()):
        values = [row[col] for col in columns]
        ws.append(values + ['', ''])

        # format row:
        for col_idx, col_name in enumerate(header, start=1):
            cell = ws.cell(row=excel_row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if col_name == 'name':
                cell.font = Font(bold=True)

        img_path = image_paths['img_path'].iloc[i]
        if os.path.exists(img_path):
            img = XLImage(img_path)
            img.height = 160
            img.width = 160
            img_cell = f"{chr(65 + len(columns) + 1)}{excel_row_idx}"
            ws.add_image(img, img_cell)

            image_paths_to_remove.append(img_path)

        excel_row_idx += 1

    wb.save(excel_path)

    for img_path in image_paths_to_remove:
        if os.path.exists(img_path):
            os.remove(img_path)
